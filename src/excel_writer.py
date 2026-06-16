from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PILImage

from models import BatchSummary, ImageTaskResult, OcrItem


HEADERS = ["序号", "原图识别内容", "玩法判定", "标准化结果", "核查截图", "金额(元)", "需人工核查", "核查原因/备注", "图片文件名"]
OLD_HEADERS = ["序号", "原图识别内容", "玩法判定", "标准化结果", "金额(元)", "需人工核查", "核查原因/备注", "图片文件名"]
PREVIOUS_HEADERS = ["序号", "原图识别内容", "玩法判定", "标准化结果", "金额(元)", "需人工核查", "核查原因/备注", "核查截图", "图片文件名"]

DETAIL_WIDTHS = [8, 26, 14, 32, 120, 12, 14, 48, 42]
REVIEW_WIDTHS = [8, 42, 30, 72, 112, 20, 20, 20]
PLAY_TYPE_SORT_ORDER = {
    "胆码": 1,
    "组三": 2,
    "组六": 3,
    "定位": 4,
    "直选": 5,
    "直选组选混合": 6,
    "未知": 99,
}


def read_existing_image_files(output_file: Path) -> set[str]:
    """读取已有Excel中“图片文件名”列，用于跳过已处理图片。"""
    if not output_file.exists():
        return set()
    try:
        wb = load_workbook(output_file, read_only=True, data_only=True)
        if "识别明细" not in wb.sheetnames:
            return set()
        ws = wb["识别明细"]
        header_map = _read_header_map(ws)
        image_col = header_map.get("图片文件名")
        if not image_col:
            return set()
        files: set[str] = set()
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=image_col).value
            if value:
                files.add(str(value).strip())
        return files
    except Exception:
        return set()
    finally:
        try:
            wb.close()
        except Exception:
            pass


def read_review_image_files(output_file: Path) -> set[str]:
    """读取已有Excel中“需人工核查=是”的图片文件名。"""
    if not output_file.exists():
        return set()
    try:
        wb = load_workbook(output_file, read_only=True, data_only=True)
        if "识别明细" not in wb.sheetnames:
            return set()
        ws = wb["识别明细"]
        header_map = _read_header_map(ws)
        image_col = header_map.get("图片文件名")
        review_col = header_map.get("需人工核查")
        if not image_col or not review_col:
            return set()
        files: set[str] = set()
        for row in range(2, ws.max_row + 1):
            image_file = ws.cell(row=row, column=image_col).value
            needs_review = ws.cell(row=row, column=review_col).value
            if image_file and str(needs_review).strip() == "是":
                files.add(str(image_file).strip())
        return files
    except Exception:
        return set()
    finally:
        try:
            wb.close()
        except Exception:
            pass


def ensure_workbook_writable(output_file: Path) -> None:
    """在发送AI前检查目标Excel能否写入，避免最后保存失败浪费调用。"""
    if not output_file.exists():
        output_file.parent.mkdir(parents=True, exist_ok=True)
        return
    try:
        with output_file.open("a+b"):
            pass
    except PermissionError as exc:
        raise PermissionError(f"目标Excel可能正在被打开，请先关闭后重试：{output_file}") from exc
    try:
        wb = load_workbook(output_file, read_only=True, data_only=True)
        if "识别明细" not in wb.sheetnames:
            wb.close()
            raise RuntimeError("缺少“识别明细”sheet")
        ws = wb["识别明细"]
        header_map = _read_header_map(ws)
        if "图片文件名" not in header_map:
            wb.close()
            raise RuntimeError("缺少“图片文件名”列")
        wb.close()
    except Exception as exc:
        raise RuntimeError(f"目标Excel无法读取或不是本程序生成的统计表：{output_file}；原因：{exc}") from exc


def write_workbook(results: list[ImageTaskResult], output_file: Path, replace_image_files: set[str] | None = None) -> Path:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    rows = _flatten_items(results)
    image_path_map = _build_image_path_map(results)
    if output_file.exists():
        return _append_workbook(results, rows, output_file, replace_image_files=replace_image_files, image_path_map=image_path_map)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "统计总览"
    ws_detail = wb.create_sheet("识别明细")
    ws_images = wb.create_sheet("截图核对")

    summary = _build_summary(results, rows)
    _write_detail_sheet(ws_detail, rows, image_path_map=image_path_map, output_file=output_file)
    _write_summary_sheet(ws_summary, results, rows, summary)
    _write_review_images_sheet(ws_images, rows, image_path_map=image_path_map, output_file=output_file)

    wb.save(output_file)
    return output_file


def upgrade_existing_workbook_layout(output_file: Path) -> None:
    """把旧版Excel升级为当前列结构；没有旧文件时不处理。"""
    if not output_file.exists():
        return
    wb = load_workbook(output_file)
    if "识别明细" in wb.sheetnames:
        ws_detail = wb["识别明细"]
        _ensure_detail_sheet_layout(ws_detail, output_file)
        _convert_images_to_move_and_size(ws_detail)
        _resort_detail_sheet_rows(ws_detail, output_file=output_file)
        _apply_image_file_hyperlinks(ws_detail, {}, output_file)
        _refresh_detail_filter_and_table(ws_detail)
    if "截图核对" in wb.sheetnames:
        ws_images = wb["截图核对"]
        _apply_review_column_widths(ws_images)
        _convert_images_to_move_and_size(ws_images)
        _apply_review_sheet_image_file_hyperlinks(ws_images, {}, output_file)
    wb.save(output_file)


def _append_workbook(
    results: list[ImageTaskResult],
    new_rows: list[OcrItem],
    output_file: Path,
    replace_image_files: set[str] | None = None,
    image_path_map: dict[str, Path] | None = None,
) -> Path:
    """已有Excel时追加新结果，保留旧明细，避免覆盖已有记录。"""
    wb = load_workbook(output_file)
    ws_detail = wb["识别明细"] if "识别明细" in wb.sheetnames else wb.create_sheet("识别明细")
    _ensure_detail_sheet_layout(ws_detail, output_file)
    _convert_images_to_move_and_size(ws_detail)

    replace_image_files = {name for name in (replace_image_files or set()) if name}
    if replace_image_files:
        _replace_detail_rows(ws_detail, new_rows, replace_image_files, image_path_map=image_path_map, output_file=output_file)
    else:
        existing_count = _count_detail_rows(ws_detail)
        _append_detail_rows(ws_detail, new_rows, start_sequence=existing_count + 1, image_path_map=image_path_map, output_file=output_file)
    _resort_detail_sheet_rows(ws_detail, image_path_map=image_path_map, output_file=output_file)
    _apply_image_file_hyperlinks(ws_detail, image_path_map or {}, output_file)
    _refresh_detail_filter_and_table(ws_detail)

    all_rows = _read_detail_rows(ws_detail)
    _rewrite_summary_sheet(wb, all_rows, results)

    ws_images = wb["截图核对"] if "截图核对" in wb.sheetnames else wb.create_sheet("截图核对")
    _ensure_review_sheet_header(ws_images)
    _apply_review_column_widths(ws_images)
    _convert_images_to_move_and_size(ws_images)
    if replace_image_files:
        _clear_review_rows_for_images(ws_images, replace_image_files)
    _append_review_items_sheet(ws_images, [item for item in new_rows if item.needs_review], image_path_map=image_path_map, output_file=output_file)
    _apply_review_sheet_image_file_hyperlinks(ws_images, image_path_map or {}, output_file)

    wb.save(output_file)
    return output_file


def _build_image_path_map(results: list[ImageTaskResult]) -> dict[str, Path]:
    return {result.image_file: result.image_path for result in results if result.image_file and result.image_path}


def _flatten_items(results: list[ImageTaskResult]) -> list[OcrItem]:
    rows: list[OcrItem] = []
    for result in sorted(results, key=lambda r: r.image_index):
        rows.extend(sorted(result.items, key=lambda item: item.item_index))
    return _sort_items_for_excel(rows)


def _sort_items_for_excel(rows: list[OcrItem]) -> list[OcrItem]:
    return sorted(
        rows,
        key=lambda item: (
            PLAY_TYPE_SORT_ORDER.get(item.play_type, 98),
            item.image_file or "",
            item.image_index,
            item.item_index,
        ),
    )


def _read_header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value is not None and str(cell.value).strip()
    }


def _reset_detail_sheet(ws) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    _write_detail_sheet(ws, [])


def _apply_detail_column_widths(ws) -> None:
    for idx, width in enumerate(DETAIL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _apply_review_column_widths(ws) -> None:
    for idx, width in enumerate(REVIEW_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _ensure_detail_sheet_layout(ws, output_file: Path) -> None:
    current = [ws.cell(row=1, column=i).value for i in range(1, max(len(HEADERS), len(OLD_HEADERS), len(PREVIOUS_HEADERS)) + 1)]
    if current[: len(HEADERS)] == HEADERS:
        _apply_detail_column_widths(ws)
        return
    if current[: len(PREVIOUS_HEADERS)] == PREVIOUS_HEADERS:
        # 兼容上一版：把“核查截图”从H列移动到“标准化结果”右侧E列。
        ws.insert_cols(5)
        for image in getattr(ws, "_images", []):
            try:
                marker = image.anchor._from
                if marker.col == 7:  # 上一版图片在H列；插入E列后逻辑列应移动到E列。
                    marker.col = 4
                    marker.rowOff = pixels_to_EMU(6)
            except Exception:
                pass
        ws.move_range(f"I1:I{ws.max_row}", rows=0, cols=-4, translate=False)
        ws.delete_cols(9)
        ws.cell(row=1, column=5, value="核查截图")
        _apply_detail_column_widths(ws)
        return
    if current[: len(OLD_HEADERS)] == OLD_HEADERS:
        # 兼容旧版：在“标准化结果”和“金额(元)”之间插入“核查截图”列。
        ws.insert_cols(5)
        ws.cell(row=1, column=5, value="核查截图")
        try:
            ws.cell(row=1, column=5).fill = ws.cell(row=1, column=4).fill.copy()
            ws.cell(row=1, column=5).font = ws.cell(row=1, column=4).font.copy()
            ws.cell(row=1, column=5).alignment = ws.cell(row=1, column=4).alignment.copy()
            ws.cell(row=1, column=5).border = ws.cell(row=1, column=4).border.copy()
        except Exception:
            pass
        for image in getattr(ws, "_images", []):
            try:
                marker = image.anchor._from
                if marker.col == 6:  # 旧版图片贴在G列“核查原因/备注”
                    marker.col = 4  # 新版移动到E列“核查截图”
                    marker.rowOff = pixels_to_EMU(6)
            except Exception:
                pass
        _apply_detail_column_widths(ws)
        return
    raise RuntimeError(f"目标Excel格式不正确，缺少标准表头，已停止追加以避免覆盖旧内容：{output_file}")


def _count_detail_rows(ws) -> int:
    if ws.max_row <= 1:
        return 0
    count = 0
    for row in range(2, ws.max_row + 1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, len(HEADERS) + 1)):
            count += 1
    return count


def _append_detail_rows(
    ws,
    rows: list[OcrItem],
    start_sequence: int,
    image_path_map: dict[str, Path] | None = None,
    output_file: Path | None = None,
) -> None:
    if not rows:
        return

    _apply_detail_column_widths(ws)

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    seq = start_sequence
    for item in rows:
        row_idx = ws.max_row + 1
        values = [
            seq,
            item.raw_text,
            item.play_type,
            item.standardized,
            "",
            item.amount,
            "是" if item.needs_review else "否",
            item.review_reason if item.needs_review else "",
            item.image_file,
        ]
        ws.append(values)
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center", vertical="center")
        _set_image_file_hyperlink(ws.cell(row=row_idx, column=9), item.image_file, image_path_map or {}, output_file)

        if item.needs_review:
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row=row_idx, column=7).font = Font(color="C00000", bold=True)
        if item.crop_path and item.crop_path.exists() or item.needs_review:
            _attach_review_image(ws, item, row_idx, col_idx=5)
        else:
            ws.row_dimensions[row_idx].height = 28
        seq += 1


def _replace_detail_rows(
    ws,
    new_rows: list[OcrItem],
    image_files: set[str],
    image_path_map: dict[str, Path] | None = None,
    output_file: Path | None = None,
) -> None:
    """删除指定图片的旧明细行，再追加新结果并重新编号。"""
    header_map = _read_header_map(ws)
    image_col = header_map.get("图片文件名")
    if not image_col:
        raise RuntimeError("识别明细缺少“图片文件名”列，无法替换旧记录")

    rows_to_delete: set[int] = set()
    for row_idx in range(ws.max_row, 1, -1):
        image_file = ws.cell(row=row_idx, column=image_col).value
        if image_file and str(image_file).strip() in image_files:
            rows_to_delete.add(row_idx)

    _remove_images_anchored_to_rows(ws, rows_to_delete)
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx, 1)

    _renumber_detail_rows(ws)
    existing_count = _count_detail_rows(ws)
    _append_detail_rows(ws, new_rows, start_sequence=existing_count + 1, image_path_map=image_path_map, output_file=output_file)


def _renumber_detail_rows(ws) -> None:
    seq = 1
    for row_idx in range(2, ws.max_row + 1):
        if any(ws.cell(row=row_idx, column=col).value not in (None, "") for col in range(2, len(HEADERS) + 1)):
            ws.cell(row=row_idx, column=1, value=seq)
            seq += 1


def _remove_images_anchored_to_rows(ws, rows: set[int]) -> None:
    if not rows:
        return
    kept_images = []
    for image in getattr(ws, "_images", []):
        try:
            row = image.anchor._from.row + 1
        except Exception:
            kept_images.append(image)
            continue
        if row not in rows:
            kept_images.append(image)
    try:
        ws._images = kept_images
    except Exception:
        pass


def _emu_to_pixels(value: int | float | None) -> int:
    if value in (None, ""):
        return 0
    try:
        return max(0, int(round(float(value) / 9525)))
    except (TypeError, ValueError):
        return 0


def _copy_anchor_marker(marker: AnchorMarker) -> AnchorMarker:
    return AnchorMarker(
        col=int(getattr(marker, "col", 0) or 0),
        colOff=int(getattr(marker, "colOff", 0) or 0),
        row=int(getattr(marker, "row", 0) or 0),
        rowOff=int(getattr(marker, "rowOff", 0) or 0),
    )


def _make_two_cell_anchor(
    row_idx: int,
    col_idx: int,
    image_width_px: int,
    image_height_px: int,
    padding_px: int = 6,
    start_marker: AnchorMarker | None = None,
) -> TwoCellAnchor:
    """让图片随单元格移动和缩放，筛选隐藏行时不会全部叠在一起。"""
    if start_marker is None:
        start = AnchorMarker(
            col=col_idx - 1,
            colOff=pixels_to_EMU(padding_px),
            row=row_idx - 1,
            rowOff=pixels_to_EMU(padding_px),
        )
    else:
        start = _copy_anchor_marker(start_marker)

    start_col_off_px = _emu_to_pixels(start.colOff)
    start_row_off_px = _emu_to_pixels(start.rowOff)
    end = AnchorMarker(
        # 截图列已经加宽，图片锚点尽量放在同一行同一列内。
        # 这样筛选隐藏该行时，图片高度会随行一起压缩，不会漂浮覆盖到可见行。
        col=start.col,
        colOff=pixels_to_EMU(start_col_off_px + max(1, int(image_width_px))),
        row=start.row,
        rowOff=pixels_to_EMU(start_row_off_px + max(1, int(image_height_px))),
    )
    return TwoCellAnchor(editAs="twoCell", _from=start, to=end)


def _add_image_to_cell(ws, image_path: Path, row_idx: int, col_idx: int, image_width: int, image_height: int) -> None:
    xl_img = XLImage(str(image_path))
    xl_img.width = image_width
    xl_img.height = image_height
    xl_img.anchor = _make_two_cell_anchor(row_idx, col_idx, image_width, image_height)
    ws.add_image(xl_img)


def _image_display_size_px(image) -> tuple[int, int]:
    width = int(round(float(getattr(image, "width", 0) or 0)))
    height = int(round(float(getattr(image, "height", 0) or 0)))
    ext = getattr(getattr(image, "anchor", None), "ext", None)
    if (width <= 0 or height <= 0) and ext is not None:
        if width <= 0:
            width = _emu_to_pixels(getattr(ext, "cx", 0))
        if height <= 0:
            height = _emu_to_pixels(getattr(ext, "cy", 0))
    return max(1, width or 160), max(1, height or 120)


def _convert_images_to_move_and_size(ws) -> None:
    """把旧版浮动截图改成“随单元格移动和缩放”，解决筛选后截图叠加。"""
    for image in getattr(ws, "_images", []):
        try:
            anchor = image.anchor
            marker = getattr(anchor, "_from", None)
            if marker is None:
                continue
            image_width, image_height = _image_display_size_px(image)
            image.anchor = _make_two_cell_anchor(
                row_idx=int(marker.row) + 1,
                col_idx=int(marker.col) + 1,
                image_width_px=image_width,
                image_height_px=image_height,
                start_marker=marker,
            )
        except Exception:
            # 锚点升级失败不影响数据本身；后续新插入图片仍会使用新锚点。
            continue


def _refresh_detail_filter_and_table(ws) -> None:
    ws.auto_filter.ref = f"A1:I{max(1, ws.max_row)}"
    try:
        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]
        if ws.max_row > 1:
            table = Table(displayName="LotteryOcrDetail", ref=f"A1:I{ws.max_row}")
            style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
    except Exception:
        # 表格样式更新失败不影响核心数据与截图。
        pass


def _read_detail_rows(ws) -> list[OcrItem]:
    header_map = _read_header_map(ws)
    rows: list[OcrItem] = []
    for row_idx in range(2, ws.max_row + 1):
        image_file = _cell_text(ws, row_idx, header_map.get("图片文件名"))
        if not image_file:
            continue
        amount_raw = ws.cell(row=row_idx, column=header_map.get("金额(元)", 6)).value
        try:
            amount = int(round(float(amount_raw or 0)))
        except (TypeError, ValueError):
            amount = 0
        rows.append(
            OcrItem(
                image_index=row_idx - 2,
                item_index=row_idx - 2,
                image_file=image_file,
                raw_text=_cell_text(ws, row_idx, header_map.get("原图识别内容")),
                play_type=_cell_text(ws, row_idx, header_map.get("玩法判定")) or "未知",
                standardized=_cell_text(ws, row_idx, header_map.get("标准化结果")),
                amount=amount,
                needs_review=_cell_text(ws, row_idx, header_map.get("需人工核查")) == "是",
                review_reason=_cell_text(ws, row_idx, header_map.get("核查原因/备注")),
            )
        )
    return rows


def _resort_detail_sheet_rows(ws, image_path_map: dict[str, Path] | None = None, output_file: Path | None = None) -> None:
    rows = _sort_items_for_excel(_read_detail_rows(ws))
    images_by_old_row = _collect_detail_row_images(ws)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    try:
        ws._images = []
    except Exception:
        pass

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, item in enumerate(rows, start=1):
        row_idx = idx + 1
        values = [
            idx,
            item.raw_text,
            item.play_type,
            item.standardized,
            "",
            item.amount,
            "是" if item.needs_review else "否",
            item.review_reason if item.needs_review else "",
            item.image_file,
        ]
        ws.append(values)
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center", vertical="center")
        _set_image_file_hyperlink(ws.cell(row=row_idx, column=9), item.image_file, image_path_map or {}, output_file)

        if item.needs_review:
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row=row_idx, column=7).font = Font(color="C00000", bold=True)

        image_key = _detail_row_identity(item)
        old_image = images_by_old_row.get(image_key)
        if old_image:
            _restore_row_image(ws, old_image, row_idx, col_idx=5)
        elif item.crop_path and item.crop_path.exists() or item.needs_review:
            _attach_review_image(ws, item, row_idx, col_idx=5)
        else:
            ws.row_dimensions[row_idx].height = 28


def _collect_detail_row_images(ws) -> dict[tuple, object]:
    result: dict[tuple, object] = {}
    for image in getattr(ws, "_images", []):
        try:
            row_idx = image.anchor._from.row + 1
            key = (
                _cell_text(ws, row_idx, 2),
                _cell_text(ws, row_idx, 3),
                _cell_text(ws, row_idx, 4),
                ws.cell(row=row_idx, column=6).value,
                _cell_text(ws, row_idx, 7),
                _cell_text(ws, row_idx, 8),
                _cell_text(ws, row_idx, 9),
            )
            result[key] = image
        except Exception:
            continue
    return result


def _detail_row_identity(item: OcrItem) -> tuple:
    return (
        item.raw_text,
        item.play_type,
        item.standardized,
        item.amount,
        "是" if item.needs_review else "否",
        item.review_reason if item.needs_review else "",
        item.image_file,
    )


def _restore_row_image(ws, old_image, row_idx: int, col_idx: int) -> None:
    try:
        image_width, image_height = _image_display_size_px(old_image)
        old_image.anchor = _make_two_cell_anchor(row_idx, col_idx, image_width, image_height)
        ws.add_image(old_image)
        ws.row_dimensions[row_idx].height = max(170, (image_height + 28) * 0.75)
    except Exception:
        pass


def _cell_text(ws, row: int, col: int | None) -> str:
    if not col:
        return ""
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value).strip()


def _rewrite_summary_sheet(wb, rows: list[OcrItem], new_results: list[ImageTaskResult]) -> None:
    if "统计总览" in wb.sheetnames:
        del wb["统计总览"]
    ws_summary = wb.create_sheet("统计总览", 0)

    grouped: dict[str, list[OcrItem]] = defaultdict(list)
    for item in rows:
        grouped[item.image_file].append(item)
    synthetic_results = [
        ImageTaskResult(
            image_index=idx,
            image_path=Path(image_file),
            image_file=image_file,
            success=True,
            items=items,
        )
        for idx, (image_file, items) in enumerate(grouped.items())
    ]
    # 保留本轮失败图片信息，便于立刻核对。
    failed_new = [r for r in new_results if not r.success]
    for failed in failed_new:
        if failed.image_file not in grouped:
            synthetic_results.append(failed)

    summary = _build_summary(synthetic_results, rows)
    _write_summary_sheet(ws_summary, synthetic_results, rows, summary)


def _ensure_review_sheet_header(ws) -> None:
    if ws.max_row < 3 or ws.cell(row=3, column=1).value != "序号":
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
        _write_review_images_sheet(ws, [])
    if ws.max_row >= 4 and ws.cell(row=4, column=1).value == "没有需要人工核查的投注行":
        ws.delete_rows(4, 1)


def _clear_review_rows_for_images(ws, image_files: set[str]) -> None:
    """从“截图核对”sheet 删除指定图片旧记录和对应旧截图。"""
    if ws.max_row < 4:
        return
    rows_to_delete: set[int] = set()
    for row_idx in range(ws.max_row, 3, -1):
        image_file = ws.cell(row=row_idx, column=2).value
        if image_file and str(image_file).strip() in image_files:
            rows_to_delete.add(row_idx)
    _remove_images_anchored_to_rows(ws, rows_to_delete)
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx, 1)
    seq = 1
    for row_idx in range(4, ws.max_row + 1):
        if ws.cell(row=row_idx, column=2).value:
            ws.cell(row=row_idx, column=1, value=seq)
            seq += 1


def _append_review_items_sheet(
    ws,
    review_items: list[OcrItem],
    image_path_map: dict[str, Path] | None = None,
    output_file: Path | None = None,
) -> None:
    if not review_items:
        return
    start_seq = 1
    for row_idx in range(4, ws.max_row + 1):
        if ws.cell(row=row_idx, column=2).value:
            start_seq += 1

    row = max(4, ws.max_row + 1)
    for offset, item in enumerate(review_items):
        seq = start_seq + offset
        ws.cell(row=row, column=1, value=seq)
        ws.cell(row=row, column=2, value=item.image_file)
        _set_image_file_hyperlink(ws.cell(row=row, column=2), item.image_file, image_path_map or {}, output_file)
        ws.cell(row=row, column=3, value=item.raw_text)
        ws.cell(row=row, column=4, value=item.review_reason)
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        if item.crop_path and item.crop_path.exists():
            with PILImage.open(item.crop_path) as img:
                image_width, image_height = img.size
            ws.row_dimensions[row].height = max(120, (image_height + 18) * 0.75)
            _add_image_to_cell(ws, item.crop_path, row, col_idx=5, image_width=image_width, image_height=image_height)
        else:
            ws.row_dimensions[row].height = 80
        row += 1


def _build_summary(results: list[ImageTaskResult], rows: list[OcrItem]) -> BatchSummary:
    return BatchSummary(
        image_count=len(results),
        success_image_count=sum(1 for r in results if r.success),
        failed_image_count=sum(1 for r in results if not r.success),
        item_count=len(rows),
        review_count=sum(1 for item in rows if item.needs_review),
        total_amount=sum(int(item.amount or 0) for item in rows),
        review_amount=sum(int(item.amount or 0) for item in rows if item.needs_review),
    )


def _write_detail_sheet(
    ws,
    rows: list[OcrItem],
    image_path_map: dict[str, Path] | None = None,
    output_file: Path | None = None,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{max(1, len(rows) + 1)}"

    _apply_detail_column_widths(ws)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 26

    for idx, item in enumerate(rows, start=1):
        row_idx = idx + 1
        values = [
            idx,
            item.raw_text,
            item.play_type,
            item.standardized,
            "",
            item.amount,
            "是" if item.needs_review else "否",
            item.review_reason if item.needs_review else "",
            item.image_file,
        ]
        ws.append(values)
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center", vertical="center")
        _set_image_file_hyperlink(ws.cell(row=row_idx, column=9), item.image_file, image_path_map or {}, output_file)

        if item.needs_review:
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row=row_idx, column=7).font = Font(color="C00000", bold=True)
        if item.crop_path and item.crop_path.exists() or item.needs_review:
            _attach_review_image(ws, item, row_idx, col_idx=5)
        else:
            ws.row_dimensions[row_idx].height = 28

    if rows:
        table = Table(displayName="LotteryOcrDetail", ref=f"A1:I{len(rows) + 1}")
        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)


def _attach_review_image(ws, item: OcrItem, row_idx: int, col_idx: int) -> None:
    text_lines = max(1, len(item.review_reason) // 34 + item.review_reason.count("\n") + 1)
    text_height_px = min(96, 20 + text_lines * 18)
    default_image_height_px = 150
    if item.crop_path and item.crop_path.exists():
        with PILImage.open(item.crop_path) as img:
            image_width, image_height = img.size
            ws.row_dimensions[row_idx].height = max(170, (max(text_height_px, image_height) + 28) * 0.75)
        _add_image_to_cell(ws, item.crop_path, row_idx, col_idx=col_idx, image_width=image_width, image_height=image_height)
    else:
        ws.row_dimensions[row_idx].height = max(120, (text_height_px + default_image_height_px) * 0.75)


def _write_summary_sheet(ws, results: list[ImageTaskResult], rows: list[OcrItem], summary: BatchSummary) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:I1")
    ws["A1"] = "投注识别统计"
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["D"].width = 24

    summary_rows = [
        ("图片总数", summary.image_count),
        ("成功图片数", summary.success_image_count),
        ("失败图片数", summary.failed_image_count),
        ("投注行数", summary.item_count),
        ("需人工核查行数", summary.review_count),
        ("金额合计(元)", summary.total_amount),
        ("需核查金额(元)", summary.review_amount),
    ]
    start = 3
    for offset, (label, value) in enumerate(summary_rows):
        row = start + offset
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="F7FBFF")

    play_stats: dict[str, list[int]] = defaultdict(lambda: [0, 0])
    for item in rows:
        play_stats[item.play_type][0] += 1
        play_stats[item.play_type][1] += int(item.amount or 0)

    ws["D3"] = "玩法判定"
    ws["E3"] = "条数"
    ws["F3"] = "金额(元)"
    for cell in ws["D3:F3"][0]:
        cell.fill = PatternFill("solid", fgColor="70AD47")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for idx, play_type in enumerate(["胆码", "组三", "组六", "定位", "直选", "直选组选混合", "未知"], start=4):
        count, amount = play_stats.get(play_type, [0, 0])
        ws.cell(row=idx, column=4, value=play_type)
        ws.cell(row=idx, column=5, value=count)
        ws.cell(row=idx, column=6, value=amount)

    ws["A13"] = "核对说明"
    ws["A13"].fill = PatternFill("solid", fgColor="F4B183")
    ws["A13"].font = Font(bold=True)
    notes = [
        "1. 红色圈注、红色合计、红色批注不计入投注内容。",
        "2. 正常行的“核查原因/备注”为空。",
        "3. 明细表“核查截图”列会尽量为每一行贴局部截图；需人工核查行会额外写明原因。",
        "4. 定位玩法禁止盲目把不确定数字转为*，触发风险时自动标记人工核查。",
    ]
    for i, note in enumerate(notes, start=14):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
        ws.cell(row=i, column=1, value=note)
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True)

    failed = [r for r in results if not r.success]
    if failed:
        ws["A20"] = "失败图片"
        ws["A20"].fill = PatternFill("solid", fgColor="C00000")
        ws["A20"].font = Font(color="FFFFFF", bold=True)
        ws["A21"] = "文件名"
        ws["B21"] = "错误"
        for i, result in enumerate(failed, start=22):
            ws.cell(row=i, column=1, value=result.image_file)
            ws.cell(row=i, column=2, value=result.error_message)


def _write_review_images_sheet(
    ws,
    rows: list[OcrItem],
    image_path_map: dict[str, Path] | None = None,
    output_file: Path | None = None,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "截图核对"
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["序号", "图片文件名", "原图识别内容", "核查原因/备注", "核查截图"]
    ws.append([])
    ws.append(headers)
    _apply_review_column_widths(ws)
    for cell in ws[3]:
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    review_items = [item for item in rows if item.needs_review]
    row = 4
    for idx, item in enumerate(review_items, start=1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item.image_file)
        _set_image_file_hyperlink(ws.cell(row=row, column=2), item.image_file, image_path_map or {}, output_file)
        ws.cell(row=row, column=3, value=item.raw_text)
        ws.cell(row=row, column=4, value=item.review_reason)
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        if item.crop_path and item.crop_path.exists():
            with PILImage.open(item.crop_path) as img:
                image_width, image_height = img.size
            ws.row_dimensions[row].height = max(170, (image_height + 28) * 0.75)
            _add_image_to_cell(ws, item.crop_path, row, col_idx=5, image_width=image_width, image_height=image_height)
        else:
            ws.row_dimensions[row].height = 80
        row += 1

    if not review_items:
        ws["A4"] = "没有需要人工核查的投注行"


def _apply_image_file_hyperlinks(ws, image_path_map: dict[str, Path], output_file: Path | None) -> None:
    header_map = _read_header_map(ws)
    image_col = header_map.get("图片文件名")
    if not image_col:
        return
    for row_idx in range(2, ws.max_row + 1):
        image_file = _cell_text(ws, row_idx, image_col)
        if image_file:
            _set_image_file_hyperlink(ws.cell(row=row_idx, column=image_col), image_file, image_path_map, output_file)


def _apply_review_sheet_image_file_hyperlinks(ws, image_path_map: dict[str, Path], output_file: Path | None) -> None:
    if ws.max_row < 4:
        return
    for row_idx in range(4, ws.max_row + 1):
        image_file = _cell_text(ws, row_idx, 2)
        if image_file:
            _set_image_file_hyperlink(ws.cell(row=row_idx, column=2), image_file, image_path_map, output_file)


def _set_image_file_hyperlink(cell, image_file: str, image_path_map: dict[str, Path], output_file: Path | None) -> None:
    image_path = _resolve_image_path_for_link(image_file, image_path_map, output_file)
    if not image_path:
        return
    # WPS 对 file:///D:/... 这类 URI 兼容性不稳定，容易提示“无法打开指定的文件”。
    # 这里使用 Windows 本地绝对路径作为 hyperlink target，例如：
    # D:\venson\投注小帮手\2026-05\xxx.jpg
    # Excel 和 WPS 通常都能直接打开。
    cell.hyperlink = str(image_path.resolve())
    cell.style = "Hyperlink"
    cell.font = Font(color="0563C1", underline="single")


def _resolve_image_path_for_link(image_file: str, image_path_map: dict[str, Path], output_file: Path | None) -> Path | None:
    if not image_file:
        return None

    mapped = image_path_map.get(image_file)
    if mapped and mapped.exists():
        return mapped

    raw_path = Path(image_file)
    if raw_path.is_absolute() and raw_path.exists():
        return raw_path

    candidates: list[Path] = []
    if output_file:
        # 默认输出结构：outputs\<图片文件夹名>\识别结果\投注识别统计_xxx.xlsx
        # 对应原图目录通常是：<项目根>\<图片文件夹名>\<图片文件名>
        try:
            if output_file.parent.name == "识别结果":
                folder_name = output_file.parent.parent.name
                root_dir = output_file.parent.parent.parent.parent
                candidates.append(root_dir / folder_name / image_file)
                candidates.append(output_file.parent.parent / image_file)
        except IndexError:
            pass
        candidates.append(output_file.parent / image_file)

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None
